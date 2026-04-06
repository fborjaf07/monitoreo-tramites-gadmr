import asyncio
import json
import os
import re
from datetime import datetime
from playwright.async_api import async_playwright
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

USUARIO = "BORJAFF"
CONTRASENA = "Amelia1010("
HORAS_ALERTA = 24
ARCHIVO_ENTRADA = "tramites_entrada.xlsx"
ARCHIVO_DASHBOARD = "dashboard_tramites.xlsx"
ARCHIVO_ALERTAS = "alertas.json"

LOGIN_URL = ("https://egob.gadmriobamba.gob.ec:8443/cas/login?service=https%3A%2F%2Fegobedoc.gadmriobamba.gob.ec%3A8081%2Fauth%2Fcas%2Fcallback%3Forigin%3Dhttps%253A%252F%252Fegobedoc.gadmriobamba.gob.ec%253A8081%252F")
BASE_URL = "https://egobedoc.gadmriobamba.gob.ec:8081/issues/"

def leer_tramites_entrada():
    if not os.path.exists(ARCHIVO_ENTRADA):
        print("ERROR: No se encuentra tramites_entrada.xlsx")
        return []
    wb = load_workbook(ARCHIVO_ENTRADA)
    ws = wb.active
    tramites = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        numero = row[0]
        descripcion = row[1] if len(row) > 1 else ""
        activo = row[2] if len(row) > 2 else "SI"
        if numero and str(numero).strip() and str(activo).upper() == "SI":
            tramites.append({
                "numero": str(numero).strip(),
                "descripcion": str(descripcion).strip() if descripcion else ""
            })
    return tramites

def calcular_tiempo(fecha_str):
    if not fecha_str:
        return None, "Desconocido"
    try:
        formatos = ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%d/%m/%Y %H:%M", "%d/%m/%Y"]
        fecha = None
        for fmt in formatos:
            try:
                fecha = datetime.strptime(str(fecha_str).strip(), fmt)
                break
            except ValueError:
                continue
        if not fecha:
            return None, "Desconocido"
        delta = datetime.now() - fecha
        horas = delta.total_seconds() / 3600
        if horas < 1:
            return horas, f"{int(delta.total_seconds()/60)} min"
        elif horas < 24:
            return horas, f"{int(horas)}h {int((horas%1)*60)}min"
        else:
            dias = int(horas // 24)
            hrs = int(horas % 24)
            return horas, f"{dias}d {hrs}h"
    except Exception:
        return None, "Desconocido"

def color_semaforo(horas):
    if horas is None:
        return "D9D9D9"
    if horas >= HORAS_ALERTA:
        return "FF4C4C"
    elif horas >= HORAS_ALERTA * 0.75:
        return "FFC000"
    else:
        return "70AD47"

async def scroll_hasta_fin(page):
    altura_anterior = 0
    for _ in range(30):
        await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        await page.wait_for_timeout(1500)
        altura_actual = await page.evaluate("document.body.scrollHeight")
        if altura_actual == altura_anterior:
            break
        altura_anterior = altura_actual

def extraer_movimientos(historico_text):
    movimientos = []
    patron_bloque = re.compile(
        r'#\d+\s*\n([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑA-Za-záéíóúñ\s\.]+?)\s*\(([^)]+)\)\s*\n(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2})[^\n]*\n(.*?)(?=#\d+\s*\n|\Z)',
        re.DOTALL
    )
    for m in patron_bloque.finditer(historico_text):
        nombre = m.group(1).strip()
        cargo = m.group(2).strip()
        fecha = m.group(3).strip()
        contenido = m.group(4)
        nota_match = re.search(r'Nota:\s*(.+?)(?:\nAsignado|\nDocumento|\nAñadido|\nEncargo|\Z)', contenido, re.DOTALL)
        nota = nota_match.group(1).strip()[:300] if nota_match else ""
        enviado_a = ""
        asig_match = re.search(r'Asignado ha cambiado de .+? a ([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s]+?)(?:\n|$)', contenido)
        if asig_match:
            enviado_a = asig_match.group(1).strip()
        else:
            env_match = re.search(r'Documento Enviado a ([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s]+?)(?:\n|$)', contenido)
            if env_match:
                enviado_a = env_match.group(1).strip()
        movimientos.append({
            "nombre": nombre,
            "cargo": cargo,
            "fecha": fecha,
            "nota": nota,
            "enviado_a": enviado_a
        })
    return movimientos

async def extraer_datos(page, numero):
    resultado = {
        "numero": numero,
        "asunto": "",
        "estado": "",
        "movimientos": [],
        "hijos": [],
        "error": None
    }
    try:
        url = BASE_URL + numero
        await page.goto(url, wait_until="networkidle")
        await page.wait_for_timeout(3000)
        await scroll_hasta_fin(page)
        texto = await page.inner_text("body")

        asunto_match = re.search(r"Asunto:\s*(.+?)(?:\n)", texto)
        if asunto_match:
            resultado["asunto"] = asunto_match.group(1).strip()

        estado_match = re.search(r"Estado:\s*(\w+)", texto)
        if estado_match:
            resultado["estado"] = estado_match.group(1).strip()

        idx_hijos = texto.find("Trámite/s hijos")
        if idx_hijos == -1:
            idx_hijos = texto.find("Tramite/s hijos")
        idx_historico = texto.find("Histórico")
        if idx_historico == -1:
            idx_historico = texto.find("Historico")

        if idx_hijos > 0 and idx_historico > idx_hijos:
            seccion_hijos = texto[idx_hijos:idx_historico]
            patron = re.compile(
                r'\(\s*\n(\d+)\s*\n\)\s*\n.*?(Nuevo|En proceso|Aprobado|Rechazado|Cerrado|Firmado|Enviado|Respuesta|Archivado)\s*\n(\d{4}-\d{2}-\d{2})',
                re.DOTALL
            )
            for match in patron.finditer(seccion_hijos):
                num_hijo = match.group(1).strip()
                estado_hijo = match.group(2).strip()
                fecha_hijo = match.group(3).strip()
                horas, tiempo_texto = calcular_tiempo(fecha_hijo)
                resultado["hijos"].append({
                    "numero_memorando": f"#{num_hijo}",
                    "estado": estado_hijo,
                    "fecha_movimiento": fecha_hijo,
                    "horas_transcurridas": horas,
                    "tiempo_texto": tiempo_texto,
                    "en_alerta": horas is not None and horas >= HORAS_ALERTA
                })

        if idx_historico > 0:
            historico_text = texto[idx_historico:]
            movimientos = extraer_movimientos(historico_text)
            # Tomar solo los ultimos 2 movimientos
            resultado["movimientos"] = movimientos[-2:] if len(movimientos) >= 2 else movimientos

    except Exception as e:
        resultado["error"] = str(e)
    return resultado

async def ejecutar():
    tramites = leer_tramites_entrada()
    if not tramites:
        return
    print(f"\n{'='*55}")
    print(f"INICIO: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Tramites a procesar: {len(tramites)}")
    print(f"{'='*55}\n")
    todos = []
    alertas = []
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            args=["--ignore-certificate-errors"]
        )
        context = await browser.new_context(ignore_https_errors=True)
        page = await context.new_page()
        print("Iniciando sesion...")
        await page.goto(LOGIN_URL)
        await page.wait_for_load_state("networkidle")
        await page.fill('input[name="username"]', USUARIO)
        await page.fill('input[name="password"]', CONTRASENA)
        await page.click('input[type="submit"], button[type="submit"]')
        await page.wait_for_load_state("networkidle")
        print("Sesion iniciada\n")
        for i, t in enumerate(tramites, 1):
            print(f"[{i}/{len(tramites)}] Tramite #{t['numero']}...")
            datos = await extraer_datos(page, t["numero"])
            datos["descripcion_entrada"] = t["descripcion"]
            todos.append(datos)
            if datos["movimientos"]:
                ultimo = datos["movimientos"][-1]
                horas_ult, _ = calcular_tiempo(ultimo["fecha"])
                if horas_ult is not None and horas_ult >= HORAS_ALERTA:
                    alertas.append({
                        "tramite": t["numero"],
                        "descripcion": t["descripcion"],
                        "quien_envio": ultimo["nombre"],
                        "responsable_actual": ultimo["enviado_a"],
                        "cargo": ultimo["cargo"],
                        "desde_cuando": ultimo["fecha"],
                        "nota": ultimo["nota"],
                        "horas": round(horas_ult, 1)
                    })
            if datos["error"]:
                print(f"   AVISO: {datos['error']}")
            else:
                movs = datos["movimientos"]
                if movs:
                    print(f"   OK | Movimientos: {len(movs)} | Ultimo: {movs[-1]['nombre']} -> {movs[-1]['enviado_a']}")
        await browser.close()
    generar_dashboard(todos)
    payload = {
        "timestamp": datetime.now().isoformat(),
        "total_alertas": len(alertas),
        "alertas": alertas,
        "mensaje_resumen": (
            f"ALERTA TRAMITES - {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
            f"Se encontraron {len(alertas)} tramite(s) sin movimiento por mas de {HORAS_ALERTA}h."
        ) if alertas else "Sin alertas activas."
    }
    with open(ARCHIVO_ALERTAS, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"\nAlertas guardadas: {ARCHIVO_ALERTAS}")
    print(f"Dashboard generado: {ARCHIVO_DASHBOARD}")
    print(f"COMPLETADO: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

def generar_dashboard(todos):
    wb = Workbook()
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick = Border(
        left=Side(style="medium", color="1F3864"),
        right=Side(style="medium", color="1F3864"),
        top=Side(style="medium", color="1F3864"),
        bottom=Side(style="medium", color="1F3864")
    )
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:J1")
    t = ws1["A1"]
    t.value = "DASHBOARD DE SEGUIMIENTO DE TRAMITES - GADMR"
    t.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill("solid", start_color="1F3864")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 38

    ws1.merge_cells("A2:J2")
    s = ws1["A2"]
    s.value = f"Ultima actualizacion: {now_str} | Direccion de Obras Publicas - GADMR"
    s.font = Font(name="Arial", size=10, color="FFFFFF", italic=True)
    s.fill = PatternFill("solid", start_color="2E4E8A")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 18

    ws1.row_dimensions[3].height = 22
    ws1.merge_cells("A3:D3")
    g1 = ws1["A3"]
    g1.value = "INFORMACION DEL TRAMITE"
    g1.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    g1.fill = PatternFill("solid", start_color="2E4E8A")
    g1.alignment = Alignment(horizontal="center", vertical="center")
    g1.border = border

    ws1.merge_cells("E3:J3")
    g2 = ws1["E3"]
    g2.value = "MOVIMIENTO"
    g2.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    g2.fill = PatternFill("solid", start_color="385723")
    g2.alignment = Alignment(horizontal="center", vertical="center")
    g2.border = border

    ws1.row_dimensions[4].height = 30
    headers = [
        "#", "N Tramite", "Descripcion", "Estado",
        "Quien Envio", "Cargo", "Enviado A", "Fecha", "Observacion", "Tiempo"
    ]
    colores = {
        1:"2E4E8A", 2:"2E4E8A", 3:"2E4E8A", 4:"2E4E8A",
        5:"4E7A2D", 6:"4E7A2D", 7:"4E7A2D", 8:"4E7A2D", 9:"4E7A2D", 10:"4E7A2D"
    }
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=4, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", start_color=colores[col])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    col_w = [4, 12, 35, 10, 28, 28, 28, 16, 40, 12]
    for i, w in enumerate(col_w, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    current_row = 5
    for idx, datos in enumerate(todos, 1):
        movimientos = datos.get("movimientos", [])
        num_filas = max(len(movimientos), 1)
        start_row = current_row

        # Merge columnas A-D para info del tramite
        if num_filas > 1:
            for col in range(1, 5):
                ws1.merge_cells(
                    start_row=start_row,
                    start_column=col,
                    end_row=start_row + num_filas - 1,
                    end_column=col
                )

        fill_tramite = "DCE6F1" if idx % 2 == 0 else "EBF3FB"

        # Info del tramite (columnas A-D)
        info_vals = [idx, datos["numero"], datos.get("descripcion_entrada",""), datos.get("estado","")]
        for col, val in enumerate(info_vals, 1):
            c = ws1.cell(row=start_row, column=col, value=val)
            c.font = Font(name="Arial", bold=True, size=9)
            c.fill = PatternFill("solid", start_color=fill_tramite)
            c.alignment = Alignment(horizontal="center" if col in [1,2,4] else "left", vertical="center", wrap_text=True)
            c.border = border

        # Filas de movimientos
        for mov_idx, mov in enumerate(movimientos):
            row = start_row + mov_idx
            ws1.row_dimensions[row].height = 45
            es_ultimo = (mov_idx == len(movimientos) - 1)
            horas, tiempo_txt = calcular_tiempo(mov["fecha"])
            en_alerta = es_ultimo and horas is not None and horas >= HORAS_ALERTA

            if es_ultimo:
                fill_mov = "FFE0E0" if en_alerta else "E2EFDA"
            else:
                fill_mov = "FFF2CC"

            enviado_a = mov.get("enviado_a", "")
            mov_vals = [
                mov["nombre"],
                mov["cargo"],
                enviado_a,
                mov["fecha"],
                mov["nota"],
                tiempo_txt if es_ultimo else ""
            ]
            for col, val in enumerate(mov_vals, 5):
                c = ws1.cell(row=row, column=col, value=val)
                c.font = Font(name="Arial", size=9,
                              bold=es_ultimo,
                              color="FF0000" if (es_ultimo and en_alerta) else "000000")
                c.fill = PatternFill("solid", start_color=fill_mov)
                c.alignment = Alignment(
                    horizontal="center" if col in [8, 10] else "left",
                    vertical="center",
                    wrap_text=True
                )
                c.border = border

            if es_ultimo:
                ws1.cell(row=row, column=10).fill = PatternFill("solid", start_color=color_semaforo(horas))

        if not movimientos:
            ws1.row_dimensions[start_row].height = 35
            c = ws1.cell(row=start_row, column=5, value="Sin movimientos registrados")
            c.font = Font(name="Arial", size=9, italic=True, color="595959")
            c.fill = PatternFill("solid", start_color="F2F2F2")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

        current_row = start_row + num_filas

    # HOJA 2: DETALLE HIJOS
    ws2 = wb.create_sheet("Detalle Hijos")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:G1")
    t2 = ws2["A1"]
    t2.value = "DETALLE DE TRAMITES HIJOS"
    t2.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t2.fill = PatternFill("solid", start_color="1F3864")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 35
    ws2.merge_cells("A2:G2")
    s2 = ws2["A2"]
    s2.value = f"Actualizado: {now_str}"
    s2.font = Font(name="Arial", size=9, color="FFFFFF", italic=True)
    s2.fill = PatternFill("solid", start_color="2E4E8A")
    s2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 16
    headers2 = ["Tramite Padre","N Memorando Hijo","Estado","Fecha Movimiento","Tiempo","Horas","Alerta"]
    ws2.row_dimensions[4].height = 28
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=4, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", start_color="2E4E8A")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    col_w2 = [14,20,14,18,14,10,10]
    for i, w in enumerate(col_w2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    row_d = 5
    for datos in todos:
        for hijo in datos["hijos"]:
            ws2.row_dimensions[row_d].height = 20
            horas = hijo["horas_transcurridas"]
            en_alerta = hijo["en_alerta"]
            fill_bg = "FFE0E0" if en_alerta else ("F2F2F2" if row_d % 2 == 0 else "FFFFFF")
            vals = [datos["numero"], hijo["numero_memorando"], hijo["estado"],
                    hijo["fecha_movimiento"], hijo["tiempo_texto"],
                    round(horas,1) if horas else "", "ALERTA" if en_alerta else "OK"]
            for col, val in enumerate(vals, 1):
                c = ws2.cell(row=row_d, column=col, value=val)
                c.font = Font(name="Arial", size=10, bold=(col==7 and en_alerta),
                              color="FF0000" if (col==7 and en_alerta) else "226600" if col==7 else "000000")
                c.fill = PatternFill("solid", start_color=fill_bg)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border
            ws2.cell(row=row_d, column=6).fill = PatternFill("solid", start_color=color_semaforo(horas))
            row_d += 1

    # HOJA 3: ALERTAS
    ws3 = wb.create_sheet("Alertas")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:I1")
    t3 = ws3["A1"]
    t3.value = "TRAMITES EN ALERTA - MAS DE 24 HORAS SIN MOVIMIENTO"
    t3.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t3.fill = PatternFill("solid", start_color="C00000")
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 35
    headers3 = ["N Tramite","Descripcion","Quien Envio","Responsable Actual","Cargo","Desde","Tiempo","Horas","Observacion"]
    ws3.row_dimensions[3].height = 28
    for col, h in enumerate(headers3, 1):
        c = ws3.cell(row=3, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", start_color="C00000")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    col_w3 = [12,35,28,28,28,16,12,8,42]
    for i, w in enumerate(col_w3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    row_a = 4
    hay_alertas = False
    for datos in todos:
        movimientos = datos.get("movimientos", [])
        if movimientos:
            ultimo = movimientos[-1]
            horas_ult, tiempo_txt = calcular_tiempo(ultimo["fecha"])
            if horas_ult is not None and horas_ult >= HORAS_ALERTA:
                hay_alertas = True
                ws3.row_dimensions[row_a].height = 45
                fill_a = "FFE0E0" if row_a % 2 == 0 else "FFF0F0"
                vals = [
                    datos["numero"], datos.get("descripcion_entrada",""),
                    ultimo["nombre"], ultimo.get("enviado_a",""),
                    ultimo["cargo"], ultimo["fecha"],
                    tiempo_txt, round(horas_ult,1) if horas_ult else "",
                    ultimo["nota"]
                ]
                for col, val in enumerate(vals, 1):
                    c = ws3.cell(row=row_a, column=col, value=val)
                    c.font = Font(name="Arial", size=10)
                    c.fill = PatternFill("solid", start_color=fill_a)
                    c.alignment = Alignment(
                        horizontal="center" if col in [1,6,7,8] else "left",
                        vertical="center", wrap_text=True
                    )
                    c.border = border
                row_a += 1
    if not hay_alertas:
        ws3.merge_cells("A4:I4")
        c = ws3.cell(row=4, column=1, value="Sin alertas activas en este momento")
        c.font = Font(name="Arial", bold=True, size=11, color="226600")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", start_color="E8F5E9")

    wb.save(ARCHIVO_DASHBOARD)
    print(f"Dashboard guardado: {ARCHIVO_DASHBOARD}")

if __name__ == "__main__":
    asyncio.run(ejecutar())